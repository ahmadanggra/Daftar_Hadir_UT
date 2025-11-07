import os
import pandas as pd
from openpyxl import load_workbook

# === Configuration ===
csv_file = os.path.join(os.getcwd(), "source.csv")
data_file = os.path.join(os.getcwd(), "source_data.xlsx")
template_path = os.path.join(os.getcwd(), "Master_Volume_Comtest.xlsx")
output_folder = os.path.join(os.getcwd(), "output")

os.makedirs(output_folder, exist_ok=True)

def process_document(template_path, output_folder, service_link, site_name, single_site, wb_src, ws_name ):
    # Load workbook using openpyxl
    wb = load_workbook(template_path)
    sheet = wb["BOQ"]
    ws_src = wb_src[ws_name]

    # Modify intended cells
    sheet["C4"] = f": {service_link}"
    sheet["C5"] = f": {site_name}"

    # Generate data
    src_range = ws_src["A2:E16"]
    target_start_col = 1
    target_start_row = 9
    for i, row in enumerate(src_range, start=0):
        for j, cell in enumerate(row, start=0):
            sheet.cell(row=target_start_row + i, column=target_start_col + j, value=cell.value)

    # Save the document under a new name
    if single_site:
        new_name = f"{service_link}.xlsx"
    else:
        new_name = f"{service_link}_{site_name}.xlsx"
    output_path = os.path.join(output_folder, new_name)

    wb.save(output_path)
    wb.close()

    print(f"✅ Saved new Excel: {output_path}")

def generate_excel():
    # open csv file as source of site name
    df = pd.read_csv(csv_file, delimiter=";", encoding="utf-8-sig")
    # Open source data workbook
    wb_src = load_workbook(data_file)

    # === MAIN LOOP ===
    for _, row in df.iterrows():
        service_link = row["service_link"].strip()
        site_name = service_link.split(" - ")
        is_single = True if len(site_name) == 1 else False
        if len(site_name) == 2:
            ws_name = site_name[0].split()[0] + "_" + site_name[0].split()[1]
            process_document(template_path, output_folder, service_link, site_name[0].split()[1], is_single, wb_src, ws_name)
            ws_name = site_name[0].split()[0] + "_" + site_name[1]
            process_document(template_path, output_folder, service_link, site_name[1], is_single, wb_src, ws_name)
        else:
            ws_name = service_link
            process_document(template_path, output_folder, service_link, service_link.split()[1], is_single, wb_src, ws_name)
    
    # Close source data workbook
    wb_src.close()

if __name__ == "__main__":
    generate_excel()
    print("🎉 All files generated successfully!")
